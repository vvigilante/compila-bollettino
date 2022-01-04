import sys,os
from openpyxl import load_workbook
from argparse import ArgumentParser
from bollettino import Bollettino
from compila import compila
from italian_int import italian_int
MAX_COL=26

def main():
    parser = ArgumentParser(description='Compila bollettino dato un file excel')
    parser.add_argument('file', type=str, nargs='+',
                        help='il file excel contenente i dati')
    parser.add_argument('--singoli', type=bool, default=False,
                        help='Genera un pdf per riga')
    parser.add_argument('--stampa', type=bool, default=True,
                        help='Genera un unico pdf per la stampa')
    args = parser.parse_args()
    data = []
    graphics = []
    for f in args.file:
        if f.endswith('.pdf'):
            graphics.append(f)
        else:
            print(f'Carico {f}...')
            wb = load_workbook(filename = f, read_only=True)
            path = os.path.dirname(f)
            mapping = {}
            ws = wb.active
            for col in range(MAX_COL):
                col_letter = chr(ord('A')+col)
                title = str(ws[f'{col_letter}1'].value).lower()
                if title.startswith('intestat'):
                    mapping['Intestatario'] = col
                elif title.startswith('caus'):
                    mapping['Causale'] = col
                elif title.startswith('ese'):
                    mapping['EseguitoDa'] = col
                elif title.startswith('via'):
                    mapping['Via'] = col
                elif title.startswith('cap'):
                    mapping['Cap'] = col
                elif title.startswith('loc'):
                    mapping['Localita'] = col
                elif title.startswith('imp'):
                    mapping['Importo'] = col
                elif title.startswith('conto'):
                    mapping['ContoNumero'] = col
                else:
                    print(f'Colonna ignorata: {title}')
                if len(mapping)==8:
                    break
            assert len(mapping)==8, "Non tutti i campi necessari sono stati trovati!"
            myrows = ws.rows
            next(myrows)
            for row in myrows:
                importo = round(row[mapping['Importo']].value*100)
                item = Bollettino(
                    row[mapping['Intestatario']].value,
                    row[mapping['Causale']].value,
                    row[mapping['EseguitoDa']].value,
                    row[mapping['Via']].value,
                    row[mapping['Cap']].value,
                    row[mapping['Localita']].value,
                    importo,
                    f'{italian_int(importo//100)}/{importo%100}',
                    row[mapping['ContoNumero']].value,
                    )
                print(item)
                data.append(item)
            wb.close()
    print("Compilo...")
    compila(data, path, args.singoli, args.stampa, graphics)

if __name__ == "__main__":
    main()